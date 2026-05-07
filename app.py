import glob
import json
import os
import re
import time
import html as _html
import streamlit as st
import requests.exceptions
from dotenv import load_dotenv
from ingest import ingest
from retriever import stream_answer, NvidiaConnectionError
from utils import check_swearing, detect_lang_label, detect_intent
from feedback import save_good_exchange, delete_exchange
from export_to_rag import sync_darija_db
from classify_comments import classify_new_rows
from suggestions import save_suggestion, load_suggestions, set_suggestion_status

load_dotenv()

# ── Clean up any leftover temp PDFs from previous crashed sessions ─────────────
for _tmp in glob.glob("temp_*.pdf"):
    try:
        os.remove(_tmp)
    except Exception:
        pass

INDEXED_FILES_PATH  = "./indexed_files.json"
CHAT_HISTORY_PATH   = "./chat_history.json"
PINNED_PDF          = os.environ.get("PINNED_PDF", "")   # e.g. tunisian_darija.pdf
DARIJA_CODE         = os.environ.get("DARIJA_CODE", "darija2025")

# ── Chat history helpers ───────────────────────────────────────────────────────
def load_chat_history() -> list:
    if os.path.exists(CHAT_HISTORY_PATH):
        try:
            with open(CHAT_HISTORY_PATH, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def save_chat_history(messages: list):
    serializable = [
        {"role": m["role"], "content": m["content"], "lang": m.get("lang", ""), "sources": []}
        for m in messages
    ]
    with open(CHAT_HISTORY_PATH, "w", encoding="utf-8") as f:
        json.dump(serializable, f, ensure_ascii=False)

def load_indexed_files():
    if os.path.exists(INDEXED_FILES_PATH):
        with open(INDEXED_FILES_PATH) as f:
            return json.load(f)
    return []

def save_indexed_files(files):
    with open(INDEXED_FILES_PATH, "w") as f:
        json.dump(files, f)

def delete_indexed_file(name: str, indexed_files: list):
    """Remove all ChromaDB chunks for a file and update the index list."""
    import chromadb
    DB_PATH = "./chroma_db"
    if os.path.exists(DB_PATH):
        try:
            client = chromadb.PersistentClient(path=DB_PATH)
            cols = client.list_collections()
            if cols:
                col = client.get_collection(cols[0].name)
                for src in (f"temp_{name}", name):
                    result = col.get(where={"source": src})
                    if result["ids"]:
                        col.delete(ids=result["ids"])
                        break
        except Exception:
            pass
    if name in indexed_files:
        indexed_files.remove(name)

def _highlight_arabizi(text: str) -> str:
    """Escape HTML, then wrap arabizi letter-numbers (3 5 7 9) in an amber span."""
    escaped = _html.escape(text)
    return re.sub(r'(?<=[a-zA-Z\'])([3579])(?=[a-zA-Z])', r'<span class="ar-num">\1</span>', escaped)

def _show_feedback_buttons(msg_idx: int, question: str, answer: str):
    """
    Thumbs up / thumbs down under each assistant message.
    👍 saves the Q&A pair to ChromaDB and stores its ID.
    👎 deletes it if it was previously saved, marks it as disliked.
    Once rated, the active button is replaced by a static coloured emoji.
    """
    fb     = st.session_state["feedback"].get(msg_idx, {})
    status = fb.get("status")

    c1, c2, _ = st.columns([1, 1, 10])

    with c1:
        if status == "liked":
            st.markdown('<span style="font-size:1.1rem;color:#22c55e;">👍</span>', unsafe_allow_html=True)
        else:
            if st.button("👍", key=f"up_{msg_idx}", help="Save as good example"):
                doc_id = save_good_exchange(question, answer)
                st.session_state["feedback"][msg_idx] = {"status": "liked", "doc_id": doc_id}
                st.toast("Saved to knowledge base")
                st.rerun()

    with c2:
        if status == "disliked":
            st.markdown('<span style="font-size:1.1rem;color:#ef4444;">👎</span>', unsafe_allow_html=True)
        else:
            if st.button("👎", key=f"down_{msg_idx}", help="Mark as bad answer"):
                if status == "liked" and fb.get("doc_id"):
                    delete_exchange(fb["doc_id"])
                    st.toast("Removed from knowledge base")
                else:
                    st.toast("Noted — won't save this exchange")
                st.session_state["feedback"][msg_idx] = {"status": "disliked", "doc_id": None}
                st.rerun()

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Ai fro 8asrit lkleb", layout="wide", initial_sidebar_state="expanded")

# ── Session state defaults ─────────────────────────────────────────────────────
for key, default in [("last_lang", "en"), ("last_intent", "none"), ("dark", False), ("feedback", {}), ("darija_auth", False)]:
    if key not in st.session_state:
        st.session_state[key] = default

# Restore chat from disk if not yet in session
if "messages" not in st.session_state:
    st.session_state["messages"] = load_chat_history()

dark = st.session_state["dark"]

# ── CSS ────────────────────────────────────────────────────────────────────────
LIGHT_MAIN = "#ffffff"
DARK_MAIN  = "#0c0c1d"
LIGHT_SUB  = "#f8faff"
DARK_SUB   = "#13132b"
DARK_ELEVATED = "#1a1a38"
LIGHT_TEXT = "#0e1117"
DARK_TEXT  = "#e8eaf6"
LIGHT_MUTED = "#9ca3af"
DARK_MUTED  = "rgba(232,234,246,0.42)"
LIGHT_BORDER = "rgba(102,126,234,0.18)"
DARK_BORDER  = "rgba(102,126,234,0.22)"

main_bg    = DARK_MAIN  if dark else LIGHT_MAIN
sub_bg     = DARK_SUB   if dark else LIGHT_SUB
elevated   = DARK_ELEVATED if dark else "#eef2ff"
text_col   = DARK_TEXT  if dark else LIGHT_TEXT
muted_col  = DARK_MUTED if dark else LIGHT_MUTED
border_col = DARK_BORDER if dark else LIGHT_BORDER
card_grad  = (f"linear-gradient(145deg, {DARK_SUB}, #151530)"
              if dark else "linear-gradient(135deg, #f8faff, #eef2ff)")
badge_bg   = ("linear-gradient(135deg,rgba(102,126,234,0.18),rgba(167,139,250,0.18))"
              if dark else "linear-gradient(135deg,rgba(102,126,234,0.12),rgba(167,139,250,0.12))")
pipe_bg    = DARK_SUB   if dark else "#ffffff"
pipe_border = "rgba(255,255,255,0.07)" if dark else "#e5e7eb"
empty_border = "rgba(255,255,255,0.09)" if dark else "#e5e7eb"

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html, body, [class*="css"] {{ font-family: 'Inter', sans-serif; }}
#MainMenu {{ visibility: hidden; }}
footer {{ visibility: hidden; }}
[data-testid="stHeader"] {{ background-color: transparent !important; }}
[data-testid="stStatusWidget"] {{ visibility: hidden; }}

/* ── Root background ── */
html, body {{
    background: {f"linear-gradient(160deg, #0c0c1d 0%, #0e0d20 60%, #0c0c1d 100%) !important" if dark else "#ffffff !important"};
    color: {text_col} !important;
}}
.stApp {{
    background: {f"linear-gradient(160deg, #0c0c1d 0%, #0e0d20 60%, #0c0c1d 100%)" if dark else "#ffffff"} !important;
    color: {text_col} !important;
}}
[data-testid="stAppViewContainer"] {{
    background: transparent !important;
    color: {text_col} !important;
}}
.main .block-container {{
    background: transparent !important;
    padding-top: 2rem;
}}

/* ── Catch-all dark: any element Streamlit renders white gets subdued ── */
{"" if not dark else """
[data-testid="stVerticalBlock"],
[data-testid="stHorizontalBlock"],
[data-testid="stMetric"],
[data-testid="stForm"],
[data-testid="column"] {
    background: transparent !important;
}
"""}

/* ── Body text ── */
p, li, span {{ color: {text_col} !important; }}
h1, h2, h3, h4, h5, h6 {{ color: {text_col} !important; }}
label {{ color: {text_col} !important; }}
.stMarkdown p {{ color: {text_col} !important; }}
div[data-testid="stText"] {{ color: {text_col} !important; }}

/* ── Chat messages ── */
[data-testid="stChatMessage"],
div[data-testid="stChatMessage"] {{
    background-color: {sub_bg} !important;
    border: 1px solid {border_col} !important;
    border-radius: 14px !important;
    margin: 6px 0 !important;
}}

/* ── Chat input area (bottom bar) ── */
[data-testid="stBottom"] {{
    background: {f"linear-gradient(0deg, #0c0c1d 70%, transparent 100%)" if dark else "linear-gradient(0deg, #ffffff 70%, transparent 100%)"} !important;
}}
[data-testid="stBottom"] > div,
[data-testid="stBottom"] > div > div {{
    background: transparent !important;
}}
[data-testid="stChatInput"],
[data-testid="stChatInput"] > div,
[data-testid="stChatInput"] > div > div {{
    background-color: {sub_bg} !important;
    border: 1px solid {border_col} !important;
    border-radius: 14px !important;
}}
[data-testid="stChatInputTextArea"],
[data-baseweb="textarea"],
[data-baseweb="base-input"] {{
    background-color: {sub_bg} !important;
    color: {text_col} !important;
}}

/* ── Text inputs & text areas ── */
.stTextInput > div > div,
.stTextInput input,
.stTextArea > div > div,
.stTextArea textarea {{
    background-color: {sub_bg} !important;
    color: {text_col} !important;
    border-color: {border_col} !important;
    border-radius: 10px !important;
}}
.stTextInput input:focus,
.stTextArea textarea:focus {{
    border-color: #667eea !important;
    box-shadow: 0 0 0 2px rgba(102,126,234,0.18) !important;
}}

/* ── Select / multiselect ── */
.stSelectbox > div > div,
.stMultiSelect > div > div {{
    background-color: {sub_bg} !important;
    border-color: {border_col} !important;
    border-radius: 10px !important;
    color: {text_col} !important;
}}
.stSelectbox [data-baseweb="select"] > div,
.stMultiSelect [data-baseweb="select"] > div {{
    background-color: {sub_bg} !important;
    color: {text_col} !important;
}}
/* Dropdown popup */
[data-baseweb="popover"] [data-baseweb="menu"],
[data-baseweb="popover"] ul {{
    background-color: {elevated} !important;
    border: 1px solid {border_col} !important;
    border-radius: 10px !important;
}}
[data-baseweb="popover"] li {{
    background-color: transparent !important;
    color: {text_col} !important;
}}
[data-baseweb="popover"] li:hover {{
    background-color: rgba(102,126,234,0.15) !important;
}}

/* ── Number / date inputs ── */
.stNumberInput > div,
.stDateInput > div {{
    background-color: {sub_bg} !important;
    border-color: {border_col} !important;
    border-radius: 10px !important;
    color: {text_col} !important;
}}

/* ── Expanders ── */
[data-testid="stExpander"] {{
    background-color: {sub_bg} !important;
    border: 1px solid {border_col} !important;
    border-radius: 12px !important;
}}
[data-testid="stExpander"] summary {{ color: {text_col} !important; }}
[data-testid="stExpander"] p {{ color: {text_col} !important; }}

/* ── Tabs ── */
[data-testid="stTabs"] [role="tablist"] {{
    background: transparent !important;
    border-bottom: 1px solid {border_col} !important;
}}
[data-testid="stTabs"] button {{ color: {text_col} !important; }}
[data-testid="stTabPanel"] {{
    background-color: transparent !important;
    padding-top: 1rem !important;
}}

/* ── Dataframe / data editor ── */
[data-testid="stDataFrame"],
[data-testid="stDataEditor"] {{
    background-color: {sub_bg} !important;
    border: 1px solid {border_col} !important;
    border-radius: 12px !important;
    overflow: hidden;
}}
[data-testid="stDataFrame"] thead th {{
    background-color: {elevated} !important;
    color: {text_col} !important;
}}
[data-testid="stDataFrame"] td {{ color: {text_col} !important; }}
/* Inner iframe for data tables */
[data-testid="stDataFrame"] iframe,
[data-testid="stDataEditor"] iframe {{
    border-radius: 10px !important;
}}

/* ── Buttons (all Streamlit variants) ── */
.stButton > button,
[data-testid="stBaseButton-secondary"],
[data-testid="stBaseButton-primary"] {{
    background: {f"linear-gradient(135deg, {sub_bg}, {elevated})" if dark else sub_bg} !important;
    color: {text_col} !important;
    border: 1px solid {border_col} !important;
    border-radius: 10px !important;
    transition: all 0.18s ease !important;
}}
.stButton > button:hover,
[data-testid="stBaseButton-secondary"]:hover,
[data-testid="stBaseButton-primary"]:hover {{
    background: {elevated} !important;
    border-color: rgba(102,126,234,0.45) !important;
    transform: translateY(-1px);
    box-shadow: 0 4px 14px rgba(102,126,234,0.18) !important;
}}

/* ── HR ── */
hr {{ border-color: {"rgba(255,255,255,0.08)" if dark else "#e5e7eb"} !important; }}

/* ── Caption / small text ── */
.stCaption, small {{ color: {muted_col} !important; }}

/* ── Alerts ── */
[data-testid="stAlert"] {{
    background-color: {sub_bg} !important;
    border: 1px solid {border_col} !important;
    border-radius: 10px !important;
}}

/* ── File uploader ── */
[data-testid="stFileUploader"] {{
    background-color: {sub_bg} !important;
    border-radius: 12px !important;
}}
[data-testid="stFileUploaderDropzone"] {{
    background-color: {sub_bg} !important;
    border: 1px dashed {border_col} !important;
    border-radius: 12px !important;
}}
[data-testid="stFileUploader"] label,
[data-testid="stFileUploaderDropzone"] span {{
    color: {text_col} !important;
}}
[data-testid="stFileUploader"] section,
[data-testid="stFileUploader"] > div {{
    background-color: transparent !important;
}}
[data-testid="stFileUploaderDropzone"] button,
[data-testid="stFileUploader"] button {{
    background-color: {sub_bg} !important;
    color: {text_col} !important;
    border: 1px solid {border_col} !important;
    border-radius: 8px !important;
}}

/* ── Progress / spinner ── */
[data-testid="stSpinner"] > div {{ border-top-color: #667eea !important; }}

/* ── Scrollbar ── */
{"" if not dark else """
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #0c0c1d; border-radius: 3px; }
::-webkit-scrollbar-thumb { background: rgba(102,126,234,0.3); border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: rgba(102,126,234,0.55); }
"""}

/* ── Sidebar ── */
section[data-testid="stSidebar"] {{
    background: linear-gradient(170deg, #0d0b25 0%, #1a1545 45%, #1c1640 75%, #141230 100%) !important;
    min-width: 260px !important;
    border-right: 1px solid rgba(102,126,234,0.12) !important;
}}
section[data-testid="stSidebar"] > div:first-child {{
    padding-top: 1.5rem;
}}
section[data-testid="stSidebar"] .stMarkdown p {{
    color: rgba(255,255,255,0.82) !important;
}}
section[data-testid="stSidebar"] .stRadio p,
section[data-testid="stSidebar"] .stRadio label,
section[data-testid="stSidebar"] .stRadio span {{
    color: rgba(255,255,255,0.88) !important;
}}
section[data-testid="stSidebar"] .stButton > button,
section[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"],
section[data-testid="stSidebar"] [data-testid="stBaseButton-primary"] {{
    background: rgba(255,255,255,0.06) !important;
    color: rgba(255,255,255,0.9) !important;
    border: 1px solid rgba(255,255,255,0.12) !important;
    width: 100%;
}}
section[data-testid="stSidebar"] .stButton > button:hover,
section[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"]:hover,
section[data-testid="stSidebar"] [data-testid="stBaseButton-primary"]:hover {{
    background: rgba(102,126,234,0.18) !important;
    border-color: rgba(102,126,234,0.4) !important;
    transform: none;
    box-shadow: none !important;
}}
section[data-testid="stSidebar"] .stCaption,
section[data-testid="stSidebar"] small {{
    color: rgba(255,255,255,0.38) !important;
}}
section[data-testid="stSidebar"] .stTextInput input,
section[data-testid="stSidebar"] .stTextArea textarea {{
    background: rgba(255,255,255,0.06) !important;
    border-color: rgba(255,255,255,0.12) !important;
    color: rgba(255,255,255,0.9) !important;
}}

/* ── Custom components ── */
.gradient-title {{
    background: linear-gradient(135deg, #667eea 0%, #a78bfa 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    font-size: 2.3rem;
    font-weight: 700;
    line-height: 1.2;
    margin-bottom: 4px;
}}
.subtitle {{
    color: {muted_col};
    font-size: 0.9rem;
    margin-bottom: 1.5rem;
}}
.section-label {{
    font-size: 0.7rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.09em;
    color: {muted_col};
    margin-bottom: 8px;
}}
.badge-row {{ display: flex; flex-wrap: wrap; gap: 6px; }}
.file-badge {{
    background: {badge_bg};
    border: 1px solid {border_col};
    border-radius: 20px;
    padding: 5px 13px;
    font-size: 0.8rem;
    color: #818cf8;
    font-weight: 500;
}}
.metric-row {{ display: flex; gap: 14px; margin: 18px 0; }}
.metric-card {{
    flex: 1;
    background: {card_grad};
    border: 1px solid {border_col};
    border-radius: 14px;
    padding: 18px 16px;
    text-align: center;
}}
.metric-card .val {{
    font-size: 2rem; font-weight: 700; color: #667eea; line-height: 1;
}}
.metric-card .lbl {{
    font-size: 0.72rem; color: {muted_col}; margin-top: 5px;
    text-transform: uppercase; letter-spacing: 0.06em;
}}
.pipe-step {{
    background: {pipe_bg};
    border: 1px solid {pipe_border};
    border-radius: 10px;
    padding: 12px 10px;
    text-align: center;
    box-shadow: {"0 2px 8px rgba(0,0,0,0.3)" if dark else "0 1px 4px rgba(0,0,0,0.06)"};
}}
.pipe-title {{ font-size: 0.8rem; font-weight: 600; color: #667eea; margin-bottom: 3px; }}
.pipe-desc  {{ font-size: 0.7rem; color: {muted_col}; }}
.empty-state {{
    text-align: center; padding: 32px 20px; font-size: 0.88rem;
    border: 1px dashed {empty_border}; border-radius: 12px; color: {muted_col};
}}
.sb-stat {{
    background: rgba(255,255,255,0.05); border-radius: 9px;
    border: 1px solid rgba(255,255,255,0.07);
    padding: 10px 14px; margin: 4px 0;
    display: flex; justify-content: space-between; align-items: center;
}}
.sb-label {{ font-size: 0.73rem; color: rgba(255,255,255,0.42); }}
.sb-value {{ font-size: 0.88rem; font-weight: 600; color: #a78bfa; }}
.detect-pill {{
    display: inline-block; border-radius: 12px; padding: 2px 10px;
    font-size: 0.75rem; font-weight: 500; margin: 2px;
}}
.pill-lang  {{ background: rgba(102,126,234,0.18); color: #818cf8; }}
.pill-intent {{ background: rgba(167,139,250,0.15); color: #a78bfa; }}
.source-row {{
    background: {sub_bg}; border-radius: 8px; padding: 10px 12px; margin: 6px 0;
    border-left: 3px solid #667eea;
}}
.source-meta {{ font-size: 0.75rem; color: {muted_col}; margin-bottom: 4px; }}

/* ── Delete button ── */
button[title^="Remove"] {{
    background: rgba(239,68,68,0.08) !important;
    color: #ef4444 !important;
    border: 1px solid rgba(239,68,68,0.22) !important;
    border-radius: 6px !important;
    padding: 2px 6px !important;
    font-size: 0.85rem !important;
    min-height: 0 !important;
}}
button[title^="Remove"]:hover {{
    background: rgba(239,68,68,0.18) !important;
    border-color: rgba(239,68,68,0.5) !important;
    transform: none !important;
    box-shadow: none !important;
}}

/* ── Feedback buttons ── */
button[title="Save as good example"],
button[title="Mark as bad answer"] {{
    background: transparent !important;
    border: 1px solid {"rgba(255,255,255,0.12)" if dark else "rgba(0,0,0,0.1)"} !important;
    border-radius: 8px !important;
    padding: 2px 7px !important;
    font-size: 1rem !important;
    min-height: 0 !important;
    line-height: 1.5 !important;
    transform: none !important;
    box-shadow: none !important;
}}
button[title="Save as good example"]:hover {{
    background: rgba(34,197,94,0.12) !important;
    border-color: #22c55e !important;
    transform: none !important;
    box-shadow: none !important;
}}
button[title="Mark as bad answer"]:hover {{
    background: rgba(239,68,68,0.12) !important;
    border-color: #ef4444 !important;
    transform: none !important;
    box-shadow: none !important;
}}

/* ── Darija response bubble ── */
.darija-tag {{
    display: inline-block;
    background: linear-gradient(135deg, rgba(245,158,11,0.14), rgba(217,119,6,0.14));
    border: 1px solid rgba(245,158,11,0.38);
    border-radius: 6px;
    padding: 2px 10px;
    font-size: 0.68rem;
    font-weight: 700;
    letter-spacing: 0.09em;
    text-transform: uppercase;
    color: #f59e0b;
    margin-bottom: 10px;
}}
.darija-bubble {{
    background: {"linear-gradient(135deg, rgba(245,158,11,0.08), rgba(217,119,6,0.05))" if dark else "linear-gradient(135deg, rgba(254,243,199,0.35), rgba(253,230,138,0.12))"};
    border: 1px solid rgba(245,158,11,0.2);
    border-left: 3px solid #f59e0b;
    border-radius: 0 12px 12px 0;
    padding: 16px 20px;
    font-size: 1.02rem;
    line-height: 2;
    letter-spacing: 0.015em;
    color: {text_col};
    word-break: break-word;
}}
.ar-num {{ color: #d97706; font-weight: 700; }}

/* ── Dark-mode HTML table (replaces dataframe iframe) ── */
.dark-table-wrapper {{
    width: 100%;
    max-height: 400px;
    overflow: auto;
    border-radius: 12px;
    border: 1px solid {border_col};
    background: {sub_bg};
    margin-bottom: 8px;
}}
.dark-table-wrapper thead th {{
    position: sticky;
    top: 0;
    z-index: 1;
}}
.dark-table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 0.84rem;
    font-family: 'Inter', sans-serif;
}}
.dark-table th {{
    background: {elevated};
    color: {muted_col};
    padding: 10px 14px;
    text-align: left;
    font-weight: 600;
    font-size: 0.75rem;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    border-bottom: 1px solid {border_col};
    white-space: nowrap;
}}
.dark-table td {{
    padding: 9px 14px;
    color: {text_col};
    border-bottom: 1px solid {"rgba(255,255,255,0.05)" if dark else border_col};
    max-width: 320px;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
}}
.dark-table tr:last-child td {{ border-bottom: none; }}
.dark-table tbody tr:hover td {{
    background: {elevated};
}}

/* ── Sidebar + main slide animation ── */
section[data-testid="stSidebar"] {{
    transition: width 0.32s cubic-bezier(0.4,0,0.2,1),
                transform 0.32s cubic-bezier(0.4,0,0.2,1) !important;
}}
.main, [data-testid="stAppViewContainer"] > section:last-child {{
    transition: margin-left 0.32s cubic-bezier(0.4,0,0.2,1) !important;
}}
</style>
""", unsafe_allow_html=True)

# ── Sidebar ────────────────────────────────────────────────────────────────────
indexed_files = load_indexed_files()

with st.sidebar:
    st.markdown("""
    <div style="padding:4px 0 20px 0; border-bottom:1px solid rgba(255,255,255,0.09); margin-bottom:16px;">
        <div style="font-size:1.5rem;font-weight:700;color:white;letter-spacing:-0.02em;">Fassarli ai</div>
        <div style="font-size:0.7rem;color:rgba(255,255,255,0.38);margin-top:2px;">RAG · Llama 3.3 · NVIDIA for 8asrit likleb</div>
    </div>
    """, unsafe_allow_html=True)

    page = st.radio("Navigate", ["Q&A", "Darija Manager", "Vector DB Explorer"], label_visibility="visible")

    st.divider()

    # Theme toggle
    theme_label = "Switch to Light Mode" if dark else "Switch to Dark Mode"
    if st.button(theme_label, use_container_width=True):
        st.session_state["dark"] = not dark
        st.rerun()

    st.divider()

    # Live detection indicators
    st.markdown("<div style='font-size:0.7rem;font-weight:600;text-transform:uppercase;letter-spacing:0.09em;color:rgba(255,255,255,0.4);margin-bottom:8px;'>Detection</div>", unsafe_allow_html=True)

    lang_display   = st.session_state.get("last_lang",   "—")
    intent_display = st.session_state.get("last_intent", "—")

    st.markdown(f"""
    <div class="sb-stat">
        <span class="sb-label">Language</span>
        <span class="sb-value">{lang_display}</span>
    </div>
    <div class="sb-stat">
        <span class="sb-label">Intent</span>
        <span class="sb-value">{intent_display}</span>
    </div>
    """, unsafe_allow_html=True)

    st.divider()

    # Stats
    st.markdown("<div style='font-size:0.7rem;font-weight:600;text-transform:uppercase;letter-spacing:0.09em;color:rgba(255,255,255,0.4);margin-bottom:8px;'>Stats</div>", unsafe_allow_html=True)
    st.markdown(f"""
    <div class="sb-stat">
        <span class="sb-label">Indexed files</span>
        <span class="sb-value">{len(indexed_files)} / 5</span>
    </div>
    """, unsafe_allow_html=True)

    st.divider()

    # Supported languages
    st.markdown("<div style='font-size:0.7rem;font-weight:600;text-transform:uppercase;letter-spacing:0.09em;color:rgba(255,255,255,0.4);margin-bottom:8px;'>Supported Languages</div>", unsafe_allow_html=True)
    st.markdown("""
    <div style="font-size:0.8rem; color:rgba(255,255,255,0.7); line-height:2;">
        🇹🇳 Tunisian Darija (Arabizi)<br>
        🇫🇷 French<br>
        🇸🇦 Arabic (MSA)<br>
        🇬🇧 English
    </div>
    <div style="margin-top:10px; font-size:0.72rem; color:rgba(255,255,255,0.35); line-height:1.5;">
        Darija is detected via keyword signals, not langdetect. Each message is analysed independently.
    </div>
    """, unsafe_allow_html=True)

    st.divider()

    st.markdown("<div style='font-size:0.7rem;font-weight:600;text-transform:uppercase;letter-spacing:0.09em;color:rgba(255,255,255,0.4);margin-bottom:8px;'>Darija Database</div>", unsafe_allow_html=True)

    if st.button("Classify New Rows", use_container_width=True, help="Send unclassified rows in darija_nlp_database.xlsx to the LLM for classification"):
        log_area = st.empty()
        log_lines: list[str] = []
        def _log(msg: str):
            log_lines.append(msg)
            log_area.code("\n".join(log_lines[-10:]))  # show last 10 lines
        with st.spinner("Classifying unvalidated rows..."):
            try:
                result = classify_new_rows(progress_callback=_log)
                if result["error"]:
                    st.error(f"Classification failed: {result['error']}")
                else:
                    st.success(
                        f"Classified {result['classified']} rows — "
                        f"{result['confident']} confident, {result['review']} need review"
                    )
            except Exception as _err:
                st.error(f"Classification failed: {_err}")

    if st.button("Sync to ChromaDB", use_container_width=True, help="Export validated rows from darija_nlp_database.xlsx and ingest into ChromaDB"):
        with st.spinner("Ingesting Darija database into ChromaDB..."):
            try:
                result = sync_darija_db()
                if result["error"]:
                    st.error(f"Sync failed: {result['error']}")
                else:
                    st.success(
                        f"Synced {result['raw_count']} intent chunks "
                        f"and {result['conv_count']} conversation chunks."
                    )
            except Exception as _sync_err:
                st.error(f"Sync failed: {_sync_err}")

# ── Q&A page ──────────────────────────────────────────────────────────────────
if page == "Q&A":
    st.markdown('<div class="gradient-title">Fassarli Ai</div>', unsafe_allow_html=True)
    st.markdown('<div class="subtitle">Ask questions in English, French, Arabic, or Tunisian Darija — hybrid retrieval + streamed answers(bilfalla9i)</div>', unsafe_allow_html=True)

    col_upload, col_files = st.columns([1.3, 1])

    with col_upload:
        uploads = st.file_uploader("Upload PDFs (max 5 total)", type="pdf", accept_multiple_files=True)
        if uploads:
            new_uploads  = [u for u in uploads if u.name not in indexed_files]
            already_done = [u for u in uploads if u.name in indexed_files]

            if len(indexed_files) + len(new_uploads) > 5:
                st.error("Maximum of 5 indexed files total.")
            else:
                for upload in already_done:
                    st.info(f"{upload.name} is already indexed.")
                for upload in new_uploads:
                    path = f"temp_{upload.name}"
                    try:
                        with open(path, "wb") as f:
                            f.write(upload.read())
                        with st.spinner(f"Indexing {upload.name}..."):
                            msg = ingest(path)
                        st.success(f"{upload.name}: {msg}")
                        indexed_files.append(upload.name)
                    finally:
                        if os.path.exists(path):
                            os.remove(path)
                if new_uploads:
                    save_indexed_files(indexed_files)

    # Auto-index pinned PDF if it exists and isn't already tracked
    if PINNED_PDF and os.path.exists(PINNED_PDF) and PINNED_PDF not in indexed_files:
        with st.spinner(f"Auto-indexing pinned file: {PINNED_PDF}..."):
            ingest(PINNED_PDF)
        indexed_files.append(PINNED_PDF)
        save_indexed_files(indexed_files)

    with col_files:
        st.markdown('<div class="section-label">Indexed files</div>', unsafe_allow_html=True)
        if indexed_files:
            for fname in list(indexed_files):
                is_pinned = (fname == PINNED_PDF)
                c1, c2 = st.columns([5, 1])
                label = f'<span class="file-badge">{fname}</span>'
                if is_pinned:
                    label += ' <span style="font-size:0.65rem;color:#f59e0b;margin-left:4px;">pinned</span>'
                c1.markdown(label, unsafe_allow_html=True)
                if is_pinned:
                    c2.markdown("🔒", unsafe_allow_html=True)
                elif c2.button("×", key=f"del_{fname}", help=f"Remove {fname} from index"):
                    with st.spinner(f"Removing {fname}..."):
                        delete_indexed_file(fname, indexed_files)
                        save_indexed_files(indexed_files)
                    st.rerun()
        else:
            st.markdown('<div class="empty-state">No files indexed yet.<br>Upload a PDF to get started.</div>', unsafe_allow_html=True)

    st.divider()

    if st.session_state.messages and st.button("Clear conversation", type="secondary"):
        st.session_state.messages = []
        st.session_state["last_intent"] = "none"
        if os.path.exists(CHAT_HISTORY_PATH):
            os.remove(CHAT_HISTORY_PATH)
        st.rerun()

    if not st.session_state.messages:
        msg = "Files are ready — ask your first question below." if indexed_files else "Upload a PDF above to start chatting."
        st.markdown(f'<div class="empty-state">{msg}</div>', unsafe_allow_html=True)

    # Render conversation history
    last_user_q = None
    for msg_idx, msg in enumerate(st.session_state.messages):
        msg_lang = msg.get("lang", "")
        with st.chat_message(msg["role"]):
            if msg["role"] == "user":
                last_user_q = msg["content"]
                st.write(msg["content"])
            elif msg["role"] == "assistant":
                if msg_lang == "darija":
                    st.markdown('<span class="darija-tag">Darija</span>', unsafe_allow_html=True)
                    content = _highlight_arabizi(msg["content"])
                    st.markdown(f'<div class="darija-bubble">{content}</div>', unsafe_allow_html=True)
                else:
                    st.write(msg["content"])
                if msg.get("sources"):
                    with st.expander("Source chunks used"):
                        for i, doc in enumerate(msg["sources"], 1):
                            score = doc.metadata.get("score")
                            relevance = f"{max(0.0, 1.0 - float(score)):.0%}" if score is not None else "BM25"
                            source = doc.metadata.get("source", "?")
                            page   = doc.metadata.get("page", "?")
                            st.markdown(f"""
                            <div class="source-row">
                                <div class="source-meta">Chunk {i} · Page {page} · {source} · Relevance: {relevance}</div>
                            </div>""", unsafe_allow_html=True)
                            st.write(doc.page_content)
                            st.divider()
                _show_feedback_buttons(msg_idx, last_user_q or "", msg["content"])

    # ── Handle new question ────────────────────────────────────────────────────
    question = st.chat_input("Ask me a question about your documents")

    # ── /improve command styling (JS watches textarea and highlights it) ───────
    import streamlit.components.v1 as _components
    _components.html("""
    <script>
    (function () {
        var CMDS = ['/improve', '/7assin', '/7assen', '/hassin'];

        function applyStyle(el, on) {
            if (on) {
                el.style.color          = '#a78bfa';
                el.style.fontWeight     = '600';
                el.style.background     = 'rgba(139,92,246,0.08)';
                el.style.borderColor    = 'rgba(139,92,246,0.55)';
                el.style.caretColor     = '#a78bfa';
                el.style.letterSpacing  = '0.015em';
                el.style.boxShadow      = '0 0 0 2px rgba(139,92,246,0.18)';
            } else {
                el.style.color         = '';
                el.style.fontWeight    = '';
                el.style.background    = '';
                el.style.borderColor   = '';
                el.style.caretColor    = '';
                el.style.letterSpacing = '';
                el.style.boxShadow     = '';
            }
        }

        function attach() {
            var ta = window.parent.document.querySelector('[data-testid="stChatInputTextArea"]');
            if (!ta) return false;
            ta.addEventListener('input', function () {
                var v = this.value.trimStart().toLowerCase();
                applyStyle(this, CMDS.some(function (c) { return v.startsWith(c); }));
            });
            return true;
        }

        if (!attach()) {
            var obs = new MutationObserver(function () { if (attach()) obs.disconnect(); });
            obs.observe(window.parent.document.body, { childList: true, subtree: true });
        }
    })();
    </script>
    """, height=0, scrolling=False)

    _IMPROVE_PREFIXES = ("/improve", "/7assin", "/7assen", "/hassin")

    if question:
        # ── /improve (and aliases) command ────────────────────────────────────
        _q_lower = question.strip().lower()
        _matched_prefix = next((p for p in _IMPROVE_PREFIXES if _q_lower.startswith(p)), None)
        if _matched_prefix:
            suggestion_text = question.strip()[len(_matched_prefix):].strip()
            if not suggestion_text:
                st.info(
                    "**Usage:** type the command followed by your correction.\n\n"
                    "Example: `/improve The definition of overfitting should mention validation loss`\n\n"
                    "Aliases: `/improve` · `/7assin` · `/7assen` · `/hassin`"
                )
            else:
                last_q = next(
                    (m["content"] for m in reversed(st.session_state.messages) if m["role"] == "user"), ""
                )
                last_a = next(
                    (m["content"] for m in reversed(st.session_state.messages) if m["role"] == "assistant"), ""
                )
                save_suggestion(last_q, last_a, suggestion_text)
                st.toast("Suggestion submitted for review — shukran! 🙏", icon="✅")
            st.stop()

        # 1. Swearing check
        is_swearing, cleaned_question = check_swearing(question)
        if is_swearing:
            st.warning("Let's keep it respectful. Continuing in 10 seconds...")
            time.sleep(10)

        # 2. Language + intent detection
        lang   = detect_lang_label(cleaned_question)
        intent = detect_intent(cleaned_question)
        st.session_state["last_lang"]   = lang
        st.session_state["last_intent"] = intent

        # 3. Auto-save previous exchange when user signals understanding
        if intent == "understood":
            prev_asst = [(i, m) for i, m in enumerate(st.session_state.messages) if m["role"] == "assistant"]
            if prev_asst:
                p_idx, p_msg = prev_asst[-1]
                if st.session_state["feedback"].get(p_idx, {}).get("status") != "liked":
                    p_user = next(
                        (m["content"] for m in reversed(st.session_state.messages[:p_idx]) if m["role"] == "user"),
                        None,
                    )
                    if p_user:
                        doc_id = save_good_exchange(p_user, p_msg["content"])
                        st.session_state["feedback"][p_idx] = {"status": "liked", "doc_id": doc_id}
                        st.toast("Good exchange saved to knowledge base")

        # 4. Last assistant answer for intent-based rephrasing
        last_answer = ""
        asst_msgs = [m for m in st.session_state.messages if m["role"] == "assistant"]
        if asst_msgs:
            last_answer = asst_msgs[-1]["content"]

        # 5. Show user message + detection pills
        st.session_state.messages.append({"role": "user", "content": question})
        with st.chat_message("user"):
            st.write(question)
            st.markdown(
                f'<div style="font-size:0.73rem;color:{muted_col};margin-top:4px;">'
                f'Detected: <span class="detect-pill pill-lang">{lang}</span> '
                f'<span class="detect-pill pill-intent">{intent}</span></div>',
                unsafe_allow_html=True,
            )

        # 6. Build conversation history (excluding latest user msg)
        history = [
            {"role": m["role"], "content": m["content"]}
            for m in st.session_state.messages[:-1]
            if m["role"] in ("user", "assistant")
        ]

        # 7. Stream answer
        with st.chat_message("assistant"):
            try:
                with st.spinner("Retrieving relevant chunks..."):
                    stream, sources = stream_answer(
                        cleaned_question,
                        chat_history=history,
                        intent=intent,
                        last_answer=last_answer,
                        lang=lang,
                    )
            except NvidiaConnectionError:
                st.error(
                    "Cannot reach the NVIDIA API — check your internet connection and try again."
                )
                st.session_state.messages.pop()  # remove the unanswered user message
                st.stop()
            if lang == "darija":
                st.markdown('<span class="darija-tag">Darija</span>', unsafe_allow_html=True)
            try:
                answer = st.write_stream(stream)
            except (
                requests.exceptions.ChunkedEncodingError,
                requests.exceptions.ConnectionError,
                requests.exceptions.ReadTimeout,
                Exception,
            ) as _stream_err:
                _err_label = type(_stream_err).__name__
                st.warning(
                    f"Connection dropped mid-response ({_err_label}). "
                    "The answer above may be incomplete — try asking again."
                )
                answer = "(Response was cut short due to a network error.)"

            if sources:
                with st.expander("Source chunks used"):
                    for i, doc in enumerate(sources, 1):
                        score     = doc.metadata.get("score")
                        relevance = f"{max(0.0, 1.0 - float(score)):.0%}" if score is not None else "BM25"
                        source    = doc.metadata.get("source", "?")
                        page      = doc.metadata.get("page", "?")
                        st.markdown(f"""
                        <div class="source-row">
                            <div class="source-meta">Chunk {i} · Page {page} · {source} · Relevance: {relevance}</div>
                        </div>""", unsafe_allow_html=True)
                        st.write(doc.page_content)
                        st.divider()

            # Feedback buttons — index = current length (position after append)
            new_idx = len(st.session_state.messages)
            _show_feedback_buttons(new_idx, cleaned_question, answer)

        st.session_state.messages.append({"role": "assistant", "content": answer, "sources": sources, "lang": lang})
        save_chat_history(st.session_state.messages)

# ── Darija Manager page ───────────────────────────────────────────────────────
elif page == "Darija Manager":
    import pandas as pd
    import openpyxl as _openpyxl

    _EXCEL = "darija_nlp_database.xlsx"
    _SHEET = "Raw Collection"
    _HEADER_ROW = 2   # Excel row 2 = pandas header index 1
    _DATA_START  = 3  # First data row in Excel

    _INTENT_OPTIONS     = ["", "confusion", "understood", "half_understood", "repeat",
                           "repeat_all", "repeat_part", "simplify", "hallucination",
                           "wrong_answer", "giving_up", "impatience", "pushback",
                           "encouragement", "other"]
    _AGGRESSION_OPTIONS = ["", "neutral", "polite", "confused", "frustrated", "aggressive", "positive"]
    _VALIDATED_OPTIONS  = ["", "yes", "review"]

    st.markdown('<div class="gradient-title">Darija Manager</div>', unsafe_allow_html=True)
    st.markdown('<div class="subtitle">Review, correct and validate LLM classifications before syncing to ChromaDB</div>', unsafe_allow_html=True)

    # ── Access code gate ──────────────────────────────────────────────────────
    if not st.session_state["darija_auth"]:
        col_gate, _ = st.columns([2, 5])
        with col_gate:
            entered = st.text_input("Access code", type="password", placeholder="Enter code to unlock")
            if st.button("Unlock", type="primary"):
                if entered == DARIJA_CODE:
                    st.session_state["darija_auth"] = True
                    st.rerun()
                else:
                    st.error("Wrong code.")
        st.stop()

    if not os.path.exists(_EXCEL):
        st.warning(f"{_EXCEL} not found in the project folder.")
        st.stop()

    df_full = pd.read_excel(_EXCEL, sheet_name=_SHEET, header=1, dtype=str)
    df_full.fillna("", inplace=True)
    df_full.columns = [str(c).strip() for c in df_full.columns]

    # Keep only rows that have a raw_comment; index = pandas index (Excel row = index + 3)
    df_data = df_full[df_full["raw_comment"].str.strip() != ""].copy()

    # ── Stats bar ────────────────────────────────────────────────────────────
    total      = len(df_data)
    n_yes      = (df_data["validated"].str.strip().str.lower() == "yes").sum()
    n_review   = (df_data["validated"].str.strip().str.lower() == "review").sum()
    n_empty    = total - n_yes - n_review

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total rows",    total)
    c2.metric("Validated",     n_yes,    delta=None)
    c3.metric("Needs review",  n_review, delta=None)
    c4.metric("Unclassified",  n_empty,  delta=None)

    st.divider()

    # ── Filter ───────────────────────────────────────────────────────────────
    col_f1, col_f2 = st.columns([2, 5])
    with col_f1:
        show_filter = st.selectbox("Show", ["All rows", "Unvalidated only", "Needs review", "Validated"])

    df_view = df_data.copy()
    if show_filter == "Unvalidated only":
        df_view = df_view[df_view["validated"].str.strip().str.lower() == ""]
    elif show_filter == "Needs review":
        df_view = df_view[df_view["validated"].str.strip().str.lower() == "review"]
    elif show_filter == "Validated":
        df_view = df_view[df_view["validated"].str.strip().str.lower() == "yes"]

    _edit_cols = ["raw_comment", "darija_arabizi", "variants", "intent_category",
                  "english_meaning", "french_meaning", "aggression_level", "validated"]
    existing = [c for c in _edit_cols if c in df_view.columns]

    tab_edit, tab_delete, tab_suggest = st.tabs(["Edit & Validate", "Select & Delete", "Suggestions"])

    # ── Tab 1: Edit ───────────────────────────────────────────────────────────
    with tab_edit:
        edited_df = st.data_editor(
            df_view[existing].copy(),
            column_config={
                "raw_comment":      st.column_config.TextColumn("Raw Comment",   disabled=True, width="large"),
                "darija_arabizi":   st.column_config.TextColumn("Arabizi",       width="medium"),
                "variants":         st.column_config.TextColumn("Variants",      width="medium"),
                "intent_category":  st.column_config.SelectboxColumn("Intent",   options=_INTENT_OPTIONS,       width="small"),
                "english_meaning":  st.column_config.TextColumn("English",       width="medium"),
                "french_meaning":   st.column_config.TextColumn("French",        width="medium"),
                "aggression_level": st.column_config.SelectboxColumn("Aggression", options=_AGGRESSION_OPTIONS, width="small"),
                "validated":        st.column_config.SelectboxColumn("Validated",  options=_VALIDATED_OPTIONS,  width="small"),
            },
            hide_index=True,
            use_container_width=True,
            num_rows="fixed",
            key="darija_editor_edit",
        )
        st.caption("Edit cells directly. Set 'validated' to 'yes' when happy, then save.")

        if st.button("Save Changes to Excel", type="primary", key="save_edits"):
            wb  = _openpyxl.load_workbook(_EXCEL)
            ws  = wb[_SHEET]
            col_idx: dict[str, int] = {}
            for col_num, cell in enumerate(ws[_HEADER_ROW], start=1):
                if cell.value:
                    col_idx[str(cell.value).strip()] = col_num
            saved = 0
            for pandas_idx, row in zip(df_view.index, edited_df.itertuples(index=False)):
                excel_row = int(pandas_idx) + _DATA_START
                for col_name in existing:
                    if col_name in col_idx:
                        ws.cell(row=excel_row, column=col_idx[col_name]).value = getattr(row, col_name, "")
                saved += 1
            wb.save(_EXCEL)
            st.success(f"Saved {saved} row(s). Click 'Sync to ChromaDB' in the sidebar when ready.")
            st.rerun()

    # ── Tab 2: Select & Delete ────────────────────────────────────────────────
    with tab_delete:
        st.caption("Click row(s) to select them, then click Delete Selected. Validated rows are protected.")

        # Only show non-validated rows here — can't delete validated ones
        df_deletable = df_view[df_view["validated"].str.strip().str.lower() != "yes"].copy()

        if df_deletable.empty:
            st.info("No unvalidated rows to delete.")
        else:
            selection = st.dataframe(
                df_deletable[existing],
                use_container_width=True,
                on_select="rerun",
                selection_mode="multi-row",
                key="darija_delete_select",
            )
            selected_positions = selection.selection.rows  # positions in df_deletable

            if selected_positions:
                selected_pandas_indices = [df_deletable.index[i] for i in selected_positions]
                st.warning(f"{len(selected_pandas_indices)} row(s) selected for deletion.")

                if st.button(f"Delete {len(selected_pandas_indices)} row(s)", type="primary", key="do_delete"):
                    wb  = _openpyxl.load_workbook(_EXCEL)
                    ws  = wb[_SHEET]
                    excel_rows = sorted(
                        [int(i) + _DATA_START for i in selected_pandas_indices],
                        reverse=True,  # delete bottom-to-top so indices don't shift
                    )
                    for excel_row in excel_rows:
                        ws.delete_rows(excel_row)
                    wb.save(_EXCEL)
                    st.success(f"Deleted {len(excel_rows)} row(s) from {_EXCEL}.")
                    st.rerun()

    # ── Tab 3: Suggestions ────────────────────────────────────────────────────
    with tab_suggest:
        st.markdown('<div class="section-label">Pending user suggestions</div>', unsafe_allow_html=True)
        st.caption("These were submitted via /improve in the chat. Approve to push to ChromaDB, reject to discard.")

        suggestions = load_suggestions()
        pending   = [s for s in suggestions if s.get("status") == "pending"]
        approved  = [s for s in suggestions if s.get("status") == "approved"]
        rejected  = [s for s in suggestions if s.get("status") == "rejected"]

        col_p, col_a, col_r = st.columns(3)
        col_p.metric("Pending",  len(pending))
        col_a.metric("Approved", len(approved))
        col_r.metric("Rejected", len(rejected))

        st.divider()

        if not pending:
            st.markdown('<div class="empty-state">No pending suggestions.</div>', unsafe_allow_html=True)
        else:
            for s in pending:
                with st.expander(f"[{s['timestamp']}]  {str(s['suggestion'])[:80]}..."):
                    st.markdown(f"**Context — student asked:**")
                    st.info(s.get("context_question") or "_(no context)_")
                    st.markdown(f"**Bot answered:**")
                    st.info(s.get("context_answer") or "_(no context)_")
                    st.markdown(f"**Suggested improvement:**")
                    st.success(s.get("suggestion", ""))

                    note = st.text_input(
                        "Reviewer note (optional)",
                        key=f"note_{s['_row']}",
                        placeholder="Why approved/rejected...",
                    )
                    ca, cr, _ = st.columns([1, 1, 4])

                    if ca.button("Approve", key=f"approve_{s['_row']}", type="primary"):
                        # Push approved suggestion to ChromaDB as a validated exchange
                        ctx_q = s.get("context_question") or "General improvement"
                        improvement = s.get("suggestion", "")
                        save_good_exchange(ctx_q, improvement)
                        set_suggestion_status(s["_row"], "approved", note)
                        st.success("Approved and added to ChromaDB.")
                        st.rerun()

                    if cr.button("Reject", key=f"reject_{s['_row']}"):
                        set_suggestion_status(s["_row"], "rejected", note)
                        st.warning("Suggestion rejected.")
                        st.rerun()

        if approved or rejected:
            st.divider()
            with st.expander(f"History ({len(approved)} approved · {len(rejected)} rejected)"):
                for s in approved + rejected:
                    status_color = "#22c55e" if s["status"] == "approved" else "#ef4444"
                    st.markdown(
                        f'<span style="color:{status_color};font-weight:600;">[{s["status"].upper()}]</span> '
                        f'`{s["timestamp"]}` — {str(s["suggestion"])[:120]}',
                        unsafe_allow_html=True,
                    )

# ── Vector DB Explorer page ────────────────────────────────────────────────────
else:
    import chromadb
    import plotly.express as px
    import pandas as pd
    import umap
    from sklearn.decomposition import PCA

    st.markdown('<div class="gradient-title">Vector DB Explorer</div>', unsafe_allow_html=True)
    st.markdown('<div class="subtitle">Inspect your indexed chunks and visualize the embedding space</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-label">Pipeline</div>', unsafe_allow_html=True)
    steps = [
        ("PDF Upload",  "PyPDFLoader → raw pages"),
        ("Chunking",    "500 chars · 50 overlap"),
        ("Embedding",   "NVIDIA nv-embedqa-e5-v5"),
        ("Storage",     "ChromaDB on disk"),
        ("Retrieval",   "BM25 + vector hybrid"),
        ("Answer",      "Llama 3.3 · streamed"),
    ]
    for col, (title, desc) in zip(st.columns(len(steps)), steps):
        col.markdown(f"""
        <div class="pipe-step">
            <div class="pipe-title">{title}</div>
            <div class="pipe-desc">{desc}</div>
        </div>""", unsafe_allow_html=True)

    st.divider()

    DB_PATH = "./chroma_db"
    if not os.path.exists(DB_PATH):
        st.warning("No `chroma_db` folder found. Upload and index a PDF in the Q&A page first.")
        st.stop()

    try:
        client      = chromadb.PersistentClient(path=DB_PATH)
        collections = client.list_collections()
        if not collections:
            st.warning("ChromaDB exists but has no collections yet.")
            st.stop()

        collection = client.get_collection(collections[0].name)
        data = collection.get(include=["documents", "metadatas", "embeddings"])
        docs, metas, vecs = data["documents"], data["metadatas"], data["embeddings"]

        if not docs:
            st.warning("Collection is empty. Index a PDF first.")
            st.stop()

        sources_set = set(m.get("source", "?") for m in metas)
        avg_len     = int(sum(len(d) for d in docs) / len(docs))

        st.markdown(f"""
        <div class="metric-row">
            <div class="metric-card"><div class="val">{len(docs)}</div><div class="lbl">Total Chunks</div></div>
            <div class="metric-card"><div class="val">{len(sources_set)}</div><div class="lbl">Source Files</div></div>
            <div class="metric-card"><div class="val">{avg_len}</div><div class="lbl">Avg Chunk Size</div></div>
        </div>""", unsafe_allow_html=True)

        st.markdown('<div class="section-label">Chunk table</div>', unsafe_allow_html=True)
        df = pd.DataFrame({
            "Chunk #":        range(1, len(docs) + 1),
            "Source":         [m.get("source", "?") for m in metas],
            "Page":           [m.get("page", "?") for m in metas],
            "Length (chars)": [len(d) for d in docs],
            "Text preview":   [d[:120] + "..." if len(d) > 120 else d for d in docs],
        })
        if dark:
            st.markdown(
                f'<div class="dark-table-wrapper">{df.to_html(index=False, border=0, classes="dark-table")}</div>',
                unsafe_allow_html=True,
            )
        else:
            st.dataframe(df, use_container_width=True, hide_index=True)

        if len(docs) < 2:
            st.info("Need at least 2 chunks to plot.")
            st.stop()

        st.markdown('<div class="section-label" style="margin-top:24px;">Embedding space</div>', unsafe_allow_html=True)
        tab1, tab2 = st.tabs(["2D UMAP", "3D PCA"])
        COLORS = px.colors.qualitative.Vivid
        plot_bg = "#474780" if dark else "#f8faff"

        with tab1:
            st.caption("Each dot is one chunk. Hover to read it. Nearby dots have similar meaning.")
            with st.spinner("Running UMAP..."):
                coords_2d = umap.UMAP(
                    n_components=2, random_state=42, n_neighbors=min(15, len(docs) - 1)
                ).fit_transform(vecs)
            fig = px.scatter(
                pd.DataFrame({
                    "x": coords_2d[:, 0], "y": coords_2d[:, 1],
                    "Chunk #": range(1, len(docs) + 1),
                    "Source": [m.get("source", "?") for m in metas],
                    "Page":   [str(m.get("page", "?")) for m in metas],
                    "Text":   [d[:200] + "..." if len(d) > 200 else d for d in docs],
                }),
                x="x", y="y", color="Source", height=550, template="plotly_white",
                hover_data={"Chunk #": True, "Page": True, "Text": True, "x": False, "y": False},
                color_discrete_sequence=COLORS,
            )
            fig.update_traces(marker=dict(size=9, opacity=0.82, line=dict(width=0.5, color="white")))
            fig.update_layout(xaxis_title="", yaxis_title="",
                              paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor=plot_bg)
            st.plotly_chart(fig, use_container_width=True)

        with tab2:
            st.caption("Drag to rotate. Closer dots have more similar meaning.")
            with st.spinner("Running PCA..."):
                coords_3d = PCA(n_components=3).fit_transform(vecs)
            fig3d = px.scatter_3d(
                pd.DataFrame({
                    "x": coords_3d[:, 0], "y": coords_3d[:, 1], "z": coords_3d[:, 2],
                    "Chunk #": range(1, len(docs) + 1),
                    "Source": [m.get("source", "?") for m in metas],
                    "Page":   [str(m.get("page", "?")) for m in metas],
                    "Text":   [d[:200] + "..." if len(d) > 200 else d for d in docs],
                }),
                x="x", y="y", z="z", color="Source", height=620, template="plotly_white",
                hover_data={"Chunk #": True, "Page": True, "Text": True, "x": False, "y": False, "z": False},
                color_discrete_sequence=COLORS,
            )
            fig3d.update_traces(marker=dict(size=5, opacity=0.85))
            fig3d.update_layout(
                paper_bgcolor="rgba(0,0,0,0)",
                scene=dict(bgcolor=plot_bg, xaxis_title="", yaxis_title="", zaxis_title=""),
            )
            st.plotly_chart(fig3d, use_container_width=True)

    except Exception as e:
        st.error(f"Could not load ChromaDB: {e}")
        st.info("Make sure you have indexed a PDF in the Q&A page first.")
