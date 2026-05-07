import json
import os
import time
import openpyxl
import _compat  # noqa: F401 — Python 3.14 + urllib3 header encoding fix
from dotenv import load_dotenv
from langchain_nvidia_ai_endpoints import ChatNVIDIA
from langchain_core.messages import SystemMessage, HumanMessage

load_dotenv()

EXCEL_PATH = "darija_nlp_database.xlsx"
SHEET_NAME = "Raw Collection"
HEADER_ROW = 2      # Row 1 is the merged title, row 2 has actual column names
DATA_START_ROW = 3

SYSTEM_PROMPT = """\
You are a Tunisian Darija NLP classifier. Your job is to classify \
Tunisian Darija comments (written in Arabizi — Latin letters mixed \
with numbers like 3=ع 5=خ 7=ح 9=ق) into structured intent data.

The user will give you a raw comment. You must respond ONLY with a \
valid JSON object, no preamble, no explanation, no markdown backticks.

Classify into exactly these fields:

{
  "darija_arabizi": "the core expression stripped of context",
  "variants": "comma separated spelling variants of the same expression",
  "intent_category": "one of: confusion, understood, half_understood, repeat, repeat_all, repeat_part, simplify, hallucination, wrong_answer, giving_up, impatience, pushback, encouragement, other",
  "english_meaning": "natural English translation",
  "french_meaning": "natural French translation",
  "aggression_level": "one of: neutral, polite, confused, frustrated, aggressive, positive",
  "validated": "yes if you are confident, review if unsure"
}

Rules:
- Tunisians skip spaces so "mafhemtch" and "ma fhemtch" are the same
- Numbers are letters: 3=ع 5=خ 7=ح 9=ق 2=ء
- Same word can have i/e and a/e vowel variations (fhemtk/fhamtk/fhimtk)
- Swear words should be noted in variants but intent still classified \
  from the rest of the sentence
- If a comment contains multiple intents pick the highest priority one:
  giving_up > hallucination > wrong_answer > confusion > half_understood \
  > simplify > repeat_all > repeat_part > repeat > impatience > understood
- If completely unrecognizable as Darija, set intent_category to "other"\
"""


def _classify_comment(llm: ChatNVIDIA, raw_comment: str) -> tuple[dict, str]:
    messages = [
        SystemMessage(content=SYSTEM_PROMPT),
        HumanMessage(content=raw_comment),
    ]
    response = llm.invoke(messages)
    raw_text = response.content.strip()
    return json.loads(raw_text), raw_text


def classify_new_rows(progress_callback=None) -> dict:
    """
    Classify unvalidated rows in the Excel file using the NVIDIA LLM.
    progress_callback(msg: str) is called for each classified row (optional).
    Returns {"classified": int, "confident": int, "review": int, "error": str|None}
    """
    llm = ChatNVIDIA(
        model="meta/llama-3.3-70b-instruct",
        temperature=0.1,
        max_tokens=300,
    )

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    col_index: dict[str, int] = {}
    for col, cell in enumerate(ws[HEADER_ROW], start=1):
        if cell.value:
            col_index[str(cell.value).strip()] = col

    required = [
        "raw_comment", "darija_arabizi", "variants", "intent_category",
        "english_meaning", "french_meaning", "aggression_level", "validated",
    ]
    missing = [c for c in required if c not in col_index]
    if missing:
        return {"classified": 0, "confident": 0, "review": 0,
                "error": f"Columns not found: {missing}"}

    if "api_raw_response" not in col_index:
        api_raw_col = ws.max_column + 1
        ws.cell(row=HEADER_ROW, column=api_raw_col, value="api_raw_response")
        col_index["api_raw_response"] = api_raw_col

    classified = confident = review_count = 0
    raw_response = ""

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        raw_comment = ws.cell(row=row_idx, column=col_index["raw_comment"]).value
        if not raw_comment:
            continue

        intent_val    = ws.cell(row=row_idx, column=col_index["intent_category"]).value
        validated_val = ws.cell(row=row_idx, column=col_index["validated"]).value
        if intent_val and str(validated_val).strip().lower() == "yes":
            continue

        try:
            result, raw_response = _classify_comment(llm, str(raw_comment))

            ws.cell(row=row_idx, column=col_index["darija_arabizi"]).value   = result.get("darija_arabizi", "")
            ws.cell(row=row_idx, column=col_index["variants"]).value          = result.get("variants", "")
            ws.cell(row=row_idx, column=col_index["intent_category"]).value   = result.get("intent_category", "")
            ws.cell(row=row_idx, column=col_index["english_meaning"]).value   = result.get("english_meaning", "")
            ws.cell(row=row_idx, column=col_index["french_meaning"]).value    = result.get("french_meaning", "")
            ws.cell(row=row_idx, column=col_index["aggression_level"]).value  = result.get("aggression_level", "")
            ws.cell(row=row_idx, column=col_index["validated"]).value         = result.get("validated", "review")

            classified += 1
            if result.get("validated") == "yes":
                confident += 1
            else:
                review_count += 1

            msg = f"Row {row_idx}: {result.get('intent_category')} — {result.get('darija_arabizi')}"
            if progress_callback:
                progress_callback(msg)
            else:
                print(f"Classified {msg}")

        except json.JSONDecodeError:
            ws.cell(row=row_idx, column=col_index["validated"]).value        = "review"
            ws.cell(row=row_idx, column=col_index["api_raw_response"]).value = raw_response
            msg = f"Row {row_idx}: invalid JSON — raw response saved"
            if progress_callback:
                progress_callback(msg)
            else:
                print(f"  {msg}")

        except Exception as exc:
            msg = f"Row {row_idx}: error — {exc}"
            if progress_callback:
                progress_callback(msg)
            else:
                print(f"  {msg}")

        time.sleep(0.5)

    wb.save(EXCEL_PATH)
    return {"classified": classified, "confident": confident,
            "review": review_count, "error": None}


def main():
    result = classify_new_rows(progress_callback=print)
    if result["error"]:
        print(f"Error: {result['error']}")
        return
    print(
        f"\nDone -- classified {result['classified']} rows. "
        f"{result['confident']} confident (yes), {result['review']} need review"
    )


if __name__ == "__main__":
    main()
