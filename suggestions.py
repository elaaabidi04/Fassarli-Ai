"""
User-submitted /improve suggestions — saved to Excel, reviewed in Darija Manager
before anything touches ChromaDB.
"""
import openpyxl
from datetime import datetime

_EXCEL  = "./darija_nlp_database.xlsx"
_SHEET  = "User Suggestions"
_COLS   = ["timestamp", "context_question", "context_answer", "suggestion", "status", "reviewer_note"]

def _ensure_sheet(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    if _SHEET not in wb.sheetnames:
        ws = wb.create_sheet(_SHEET)
        ws.append(_COLS)
    return wb[_SHEET]

def save_suggestion(context_question: str, context_answer: str, suggestion: str) -> None:
    wb = openpyxl.load_workbook(_EXCEL)
    ws = _ensure_sheet(wb)
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        (context_question or "")[:600],
        (context_answer   or "")[:600],
        suggestion,
        "pending",
        "",
    ])
    wb.save(_EXCEL)

def load_suggestions() -> list[dict]:
    wb = openpyxl.load_workbook(_EXCEL)
    ws = _ensure_sheet(wb)
    wb.save(_EXCEL)
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if any(cell is not None for cell in row):
            rows.append({"_row": i, **dict(zip(_COLS, row))})
    return rows

def set_suggestion_status(row_idx: int, status: str, note: str = "") -> None:
    wb = openpyxl.load_workbook(_EXCEL)
    ws = _ensure_sheet(wb)
    ws.cell(row=row_idx, column=5, value=status)
    ws.cell(row=row_idx, column=6, value=note)
    wb.save(_EXCEL)
